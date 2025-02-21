import pandas as pd
import matplotlib.pyplot as plt
import re
from openpyxl import load_workbook
from matplotlib import font_manager as fm
import numpy as np
import matplotlib.lines as mlines
import pdb
file_path = "mt_ah.xlsx"

# 加载自定义字体
simhei_font_path = "../SimHei.ttf"
simhei_font = fm.FontProperties(fname=simhei_font_path)

# 三类属性
difficulty_set = {"diff_bot", "diff_top", "diff_bot_group", "diff_top_group"}
separability_set = {"sep_bot", "sep_top", "sep_bot_group", "sep_top_group"}
stability_set = {"stab_bot", "stab_top", "stab_bot_group", "stab_top_group"}
our_baseline = difficulty_set | separability_set | stability_set
group_set = {"diff_bot_group", "diff_top_group", "sep_bot_group", "sep_top_group", "stab_bot_group", "stab_top_group", "random_group"}

# 读取 Excel
wb = load_workbook(file_path, read_only=True)
sheet_names = wb.sheetnames[:4]  # 只处理前 4 个 Sheet

# **1️⃣ 调整 figsize 让子图严格保持 3:2**
fig, axes = plt.subplots(2, 2, figsize=(12, 8))

for j, sheet_name in enumerate(sheet_names):
    ax = axes[j // 2, j % 2]
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # 查找 mtbench 和 arena-hard 所在行
    mtbench_row = df[df[0].astype(str).str.contains('mtbench', na=False)].index[0]
    arena_row = df[df[0].astype(str).str.contains('arena-hard', na=False)].index[0]

    # 处理 X 轴（MTBench 分数）
    x_values = df.iloc[mtbench_row, 1:].dropna().astype(float).tolist()

    # 处理 Y 轴（Arena-hard 分数）
    y_values = []
    pattern = r"score:\s*([+-]?\d+\.?\d*)"
    for cell in df.iloc[arena_row, 1:]:
        if pd.isna(cell):
            continue
        match = re.search(pattern, str(cell))
        if match:
            y_values.append(float(match.group(1)))

    if len(x_values) != len(y_values):
        print(f"⚠️ 警告: Sheet '{sheet_name}' 数据长度不匹配，跳过！")
        continue

    # 计算坐标范围
    x_min, x_max = min(x_values), max(x_values)
    y_min, y_max = min(y_values), max(y_values)
    x_margin = 0.1 * (x_max - x_min)
    y_margin = 0.1 * (y_max - y_min)
    x_min -= x_margin
    x_max += x_margin
    y_min -= y_margin
    y_max += y_margin

    ax.set_title(f"{sheet_name}", fontproperties=simhei_font, fontsize=16, fontweight='bold')
    ax.set_xlabel("MTBench Score", fontproperties=simhei_font, fontsize=14, fontweight='bold')
    ax.set_ylabel("Arena-hard Score", fontproperties=simhei_font, fontsize=14, fontweight='bold')

    # 读取列名
    columns = df.iloc[0, 1:].dropna().tolist()
    base_coords, random_coords = None, None

    # print(sheet_name)
    # print(columns)
    # pdb.set_trace()
    for i, (x, y) in enumerate(zip(x_values, y_values)):
        col_name = columns[i]

        # 过滤不需要的点
        if col_name.endswith("_bot") or col_name.endswith("_bot_group") or col_name.endswith("_bottom"):
            continue
        if col_name == 'random_group':
            continue

        # 颜色设置
        color = "#BE8FC7"  # 默认 Baseline
        if col_name in difficulty_set:
            color = "#8AC291"
        elif col_name in separability_set:
            color = "#F69A75"
        elif col_name in stability_set:
            color = "#8AB2F9"

        # 形状设置
        marker = "s"
        if col_name in group_set:
            marker = "*"
        elif col_name in our_baseline:
            marker = "o"

        # 绘制点
        if col_name == "Multi.":
            ax.scatter(x, y, color="#FF0000", marker="*", s=150, alpha=0.7, edgecolors="red")
        elif marker=='s':
            ax.scatter(x, y, color=color, marker=marker, s=50, alpha=0.7, edgecolors="black", linewidths=0.5)
        else:
            ax.scatter(x, y, color=color, marker=marker, s=70, alpha=0.7, edgecolors="black", linewidths=0.5)

        if color == "#BE8FC7":
            label_text = col_name.split("_")[0]
            ax.annotate(label_text, xy=(x, y), xytext=(5, 5),
                         textcoords='offset points', fontsize=8)

        # 记录 base 和 random 坐标
        if col_name == "Base":
            base_coords = (x, y)
        elif col_name == "Random":
            random_coords = (x, y)

    # **2️⃣ 绘制 L1/L2 曲线 + 颜色填充**
    if base_coords and random_coords:
        x_base, y_base = base_coords
        x_random, y_random = random_coords

        curve_x = np.linspace(x_min, x_max, 400)
        l1_y = np.sqrt(np.maximum(0, 10 * (x_random - x_min + 0.6) ** 2 + (y_random - y_min) ** 2 - 10 * (curve_x - x_min + 0.6) ** 2)) + y_min
        l2_y = np.sqrt(np.maximum(0, 10 * (x_base - x_min + 0.4) ** 2 + (y_base - y_min) ** 2 - 10 * (curve_x - x_min + 0.4) ** 2)) + y_min

        ax.fill_between(curve_x, 0, l2_y, color='#CAD4EB', alpha=0.1)
        ax.fill_between(curve_x, l2_y, l1_y, color='#EAF3E2', alpha=0.1)
        ax.fill_between(curve_x, l1_y, y_max, color='#F8D3D2', alpha=0.1)

        ax.plot(curve_x, l1_y, linestyle="--", color="#FEC48F", label="l1")
        ax.plot(curve_x, l2_y, linestyle="--", color="#FFE3C2", label="l2")

    # **3️⃣ X 轴刻度减少一半**
    ax.set_xticks(ax.get_xticks()[::2])
    # 计算 Y 轴刻度
    y_ticks = np.linspace(y_min, y_max, num=5)  # 生成 5 个刻度
    y_ticks = np.round(y_ticks, 1)  # 保留 1 位小数

    # 强制 Matplotlib 只使用这 5 个刻度
    ax.set_yticks(y_ticks)
    ax.set_yticklabels([f"{tick:.1f}" for tick in y_ticks], fontsize=12)  # 统一格式 & 增加字体大小

    # 设置 X/Y 轴范围
    ax.set_xlim(x_min, x_max)
    ax.set_ylim(y_min, y_max)

# **4️⃣ 设置全局图例放到下面**
fig.legend(handles=[
    mlines.Line2D([], [], color='#F69A75', linestyle='-', linewidth=2, label='Separability'),
    mlines.Line2D([], [], color='#8AB2F9', linestyle='-', linewidth=2, label='Stability'),
    mlines.Line2D([], [], color='#8AC291', linestyle='-', linewidth=2, label='Difficulty'),
    mlines.Line2D([], [], color='#BE8FC7', linestyle='-', linewidth=2, label='Baseline'),
    mlines.Line2D([], [], color='black', marker='o', markersize=10, markerfacecolor='none', linestyle='None', label='w.o. cluster'),
    mlines.Line2D([], [], color='black', marker='*', markersize=12, markerfacecolor='none', linestyle='None', label='w. cluster')
], loc='lower center', fontsize=12, frameon=True, ncol=6, bbox_to_anchor=(0.5, -0.001))

plt.tight_layout(rect=[0, 0.05, 1, 1])
plt.savefig("different_models_on_fft.jpg", dpi=300, bbox_inches='tight')
plt.savefig("different_models_on_fft.pdf", dpi=300, bbox_inches='tight')
plt.show()

print("✅ 所有图表已生成！")
